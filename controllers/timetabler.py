# -*- coding: utf-8 -*-
from timetabler_functions import (module_markdown, update_module_record_with_dates,
                                  update_event_record_with_dates, convert_date_to_weekdaytime)
import itertools
import io
import gluon
import datetime
from dateutil import parser
import pypandoc
import openpyxl
from tempfile import NamedTemporaryFile
import markdown # gluon provides MARKDOWN but lacks extensions.
import os


@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def overview():

    return dict()


## TABLE VIEW CONTROLLERS 
## - These can all be viewed by anyone but you have to log in to edit
##   and many are locked down to admin staff only

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def locations_table():
    """Presents a view of teaching staff. 
    
    Only Admins can edit, delete or add
    """
    
    db.locations.id.readable = False
    
    is_admin = auth.has_membership('admin')
    
    form = SQLFORM.grid(db.locations,
                        editable=is_admin,
                        deletable=is_admin,
                        create=is_admin, 
                        csv=is_admin)
    
    return dict(form=form)

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def courses_table():
    """Presents a view of courses. Only admins can edit, delete or add
    """
    
    db.courses.id.readable = False
    
    is_admin = auth.has_membership('admin')
    
    form = SQLFORM.grid(db.courses,
                        ignore_common_filters=is_admin,
                        editable=is_admin,
                            deletable=is_admin,
                        create=is_admin, 
                        csv=is_admin)
    
    return dict(form=form)

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def college_dates_table():
    """Presents a view of college dates. Only admins can edit or add. These are 
    not deletable because some timetabling features rely on key dates in here.
    """
    
    db.college_dates.id.readable = False
    
    is_admin = auth.has_membership('admin')
    
    form = SQLFORM.grid(db.college_dates,
                        editable=is_admin,
                        details=False,
                        deletable=False,
                        create=is_admin, 
                        csv=is_admin)
    
    # Clear the FIRST DAY cache in case that has been updated
    cache.ram('FIRST_DAY', None)
    
    return dict(form=form)

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def recurring_events_table():
    """Presents a view of recurring events. Only admins can edit, delete or add.
    """
    
    db.college_dates.id.readable = False
    
    is_admin = auth.has_membership('admin')
    
    form = SQLFORM.grid(db.recurring_events,
                        editable=is_admin,
                        deletable=is_admin,
                        create=is_admin, 
                        csv=is_admin)
    
    return dict(form=form)

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def events_table():
    """Presents a view of module events. 
    
    Timetablers should always edit events via events() but admins can access
    more details via this interface
    """
    
    db.events.id.readable = False
    is_admin = auth.has_membership('admin')
    
    form = SQLFORM.grid(db.events,
                        fields=[db.events.module_id,
                                db.events.title,
                                db.events.teacher_id],
                        editable=is_admin,
                        deletable=False,
                        create=False,
                        ignore_common_filters=is_admin,
                        csv=is_admin)
    
    return dict(form=form)

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def modules_table():
    """Presents a view of modules. These should be edited via module_information()
    but admins can create, delete and edit them here. 
    TODO - enable fullcalendar interface to changing time for admins
    """
    
    db.modules.id.readable = False
    
    is_admin = auth.has_membership('admin')
    
    form = SQLFORM.grid(db.modules,
                        fields=[db.modules.title,
                                db.modules.convenors,
                                db.modules.courses],
                        editable=is_admin,
                        deletable=is_admin,
                        create=is_admin,
                        ignore_common_filters=is_admin,
                        csv=is_admin)
    
    return dict(form=form)

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def archive_timetable():
    """A controller to push out an archive file of the timetabler data.
    """
    
    
    timetabler_tables = ['locations', 'teaching_staff', 'courses', 'modules', 
                         'events', 'college_dates', 'recurring_events']
    
    outfile = io.StringIO()
    
    for tbl in timetabler_tables:
        
        outfile.write(f"TABLE {tbl}\n")
        rows = db(db[tbl]).select()
        rows.export_to_csv_file(outfile)
    
    raise HTTP(200, outfile.getvalue(),
               **{'Content-Type': 'text/plain',
                  'Content-Disposition': 'attachment; filename=Archived_timetabler_tables.csv'})


## MODULE INFORMATION CONTROLLERS
## - information shows a form for the module level data
## - events provides a week calendar for event editing
## - view renders the existing information as html
## - doc downloads the existing information as docx or latex

def _module_page_header(module_id, current):
    """A helper function to generate the common header for module controllers.
    It also embeds some module data in the html for client side use by the 
    events calendar
    """
    
    # Set of links to create - keep the current link as plain text
    links = [{'title': 'Info', 'url': URL('module_information', args=[module_id])},
             {'title': 'Events', 'url': URL('module_events', args=[module_id])},
             {'title': 'HTML', 'url': URL('module_view', args=[module_id])},
             {'title': 'Word', 'url': URL('module_doc', args=[module_id, 'docx'])},
             {'title': 'LaTeX', 'url': URL('module_doc', args=[module_id, 'latex'])}]
    links = [SPAN(lnk['title'], _style="margin:5px")
                if lnk['title'] == current 
                else SPAN(A(lnk['title'], _href=lnk['url']), _style="margin:5px")
                for lnk in links]
    
    module_record = db.modules[module_id]
    update_module_record_with_dates(module_record)
    module_header = DIV(DIV(H2(module_record.title),
                            _class='col-8'),
                        DIV(B(CAT(links[0], ' + ', links[1], ' = ', links[2] ,
                                  XML('&or;'), links[3], XML('&or;'), links[4], 
                                  _class='pull-right')),
                            _class='col-4'),
                        _class='row', _id='module_data', 
                        _module_id = module_id, _module_start = module_record.start) + HR()
    
    return module_header

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def module_information():
    """Controller to return a SQLFORM for module level information"""
    module_id = request.args[0]
    
    form = SQLFORM(db.modules, 
                   fields = ['title',
                             'convenors',
                             'courses',
                             'placeholder_week',
                             'placeholder_n_weeks',
                             'description',
                             'aims',
                             'reading',
                             'delivery',
                             'other_notes'],
                   showid=False,
                   record=module_id, 
                   readonly=not auth.has_membership('timetabler'))
    
    if form.process().accepted:
        redirect(URL('module_view', args=[module_id]))
    
    return dict(form=form, module_data=_module_page_header(module_id, 'Info'))

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def module_events():
    """Controller to deliver a module level calendar and handle event creation and update.
    Editing is only possible for logged in users, but anyone can view and click through."""
    
    if len(request.args) == 0:
        redirect(URL('module_grid'))
    
    try:
        module_id = int(request.args[0])
    except ValueError:
        session.flash = 'Invalid module id'
        redirect(URL('module_grid'))
    
    module_record = db.modules[module_id]
    if module_record is None:
        session.flash = 'Unknown module id'
        redirect(URL('module_grid'))
        
    update_module_record_with_dates(module_record)
    
    if auth.has_membership('timetabler'):
        db.events._common_filter = None
    
    if len(request.args) == 2:
        try:
            event_id = int(request.args[1])
        except ValueError:
            session.flash = 'Invalid event id'
            redirect(URL('module_events', args=module_id))
        
        event_record = db.events[event_id]
        
        if event_record is None:
            session.flash = 'Unknown event id'
            redirect(URL('module_events', args=module_id))
        
        update_event_record_with_dates(event_record)
        event_data = DIV(_id='event_data', _event_id=event_id,
                         _start = event_record.start)
    else:
        event_id = None
        event_data = DIV()
    
    # USE SQLFORM to create the event form when an event id is passed in the arguments
    # otherwise create some instructions. The day, time and duration fields are not 
    # included. This information is stored in hidden fields in the form and updated 
    # by using fullcalendar - the resulting information is copied back into the form
    # internally below
    
    if event_id is not None:
        if event_record.conceal:
            conceal_toggle = BUTTON('Reveal', _type="submit", 
                                     _class='btn btn-secondary', _name='reveal')
        else:
            conceal_toggle = BUTTON('Conceal', _type="submit", 
                                     _class='btn btn-secondary', _name='conceal')
            
        form = SQLFORM(db.events,
                       record = event_record,
                       fields = ['title', 'description', 'teacher_id', 
                                 'location_id', 'courses'],
                       hidden = dict(start=event_record.start,
                                     duration=event_record.duration),
                       #formstyle='bootstrap3_stacked',
                       showid=False,
                       readonly=not auth.has_membership('timetabler'),
                       buttons=[A(BUTTON('Back', _class='btn btn-secondary'),
                                  _href=URL('module_events', args=module_id)),
                                BUTTON('Submit', _type="submit", 
                                       _class='btn btn-secondary', _name='submit'),
                                conceal_toggle,
                                BUTTON('Duplicate', _type="submit",
                                       _class='btn btn-secondary', _name='duplicate'),
                                BUTTON('Delete', _type="submit",  
                                       _class='btn btn-secondary', _name='delete')])
        
        if auth.has_membership('timetabler'):
            # Shrink the description box and let the buttons use the full row 
            form.element('[name=description]')['_rows']= 4
            form.element('[name=teacher_id]')['_size']= 4
            form.element('[name=location_id]')['_size']= 4
            form.element('[name=courses]')['_size']= 4
            form.element('[id=submit_record__row]').elements()[1]['_class'] = 'col-sm-12'
            # Add a name to the form, making it easier to target client side for 
            # experimenting with AJAX updating of the events window. This would need
            # a lot of reworking to move it all client side.
            
            # form.element()['_name'] = 'event_details'
            # form.element()['_action'] = ''
    
    elif auth.has_membership('timetabler'):
        form= DIV(H4('Options'),
                  UL(LI('Click on an existing event to unlock it for editing. '
                        'You can drag and resize the event to reschedule it and '
                        'edit the event details in the form that will appear. '
                        'Remember to press the Submit button after making changes!',
                        #_style='padding:2px'
                        ),
                     LI('Drag and drop the event below to create a new event:',
                        CENTER(DIV(DIV(DIV(CENTER('New event!'),
                                    _class='fc-event-main', 
                                    **{'data-event': '{ "title": "newevent", "duration": "01:00" }'}),
                                _class='fc-event fc-h-event fc-daygrid-event fc-daygrid-block-event',
                                _style='padding:10px;width:150px'),
                            _id='external-events')),
                        _style='padding:2px'),
                     LI('You can use ', 
                        A('Markdown', _href='https://guides.github.com/features/mastering-markdown/'),
                        ' to easily add formatting to your content.',
                        _style='padding:2px'),
                     LI('Follow these links to add new ', 
                        A('teaching staff', _href=URL('teaching_staff')),
                        ' or ',
                        A('locations', _href=URL('locations')),
                        ' if needed.',
                        _style='padding:2px'),
                     LI('Once selected, events can also be:', 
                          UL(LI(B('Duplicated'), ' to make an exact copy for editing'),
                             LI(B('Concealed'), ' to retain an event in the database '
                                  'but hide it from public view'),
                             LI(B('Deleted'), ' to permanently remove the event.')))),
                  _class='jumbotron')
    else:
        form= DIV(H4('Options'),
                  UL(LI('Log in to edit events.'),
                     LI('Click on an event to see event details.')),
                  _class='jumbotron')
    
    # Process the form if that is what comes back, first adding the hidden fields back in
    if isinstance(form, gluon.sqlhtml.SQLFORM):
        form.vars.start = request.vars.start
        form.vars.duration = request.vars.duration
    
    if isinstance(form, gluon.sqlhtml.SQLFORM) and form.process().accepted:
        # The request variables contain the name of the submit button,
        # which is used here to identify the action.
        
        if 'conceal' in request.vars:
            event_record.update_record(conceal=True)
        elif 'reveal' in request.vars:
            event_record.update_record(conceal=False)
        elif 'delete' in request.vars:
            event_record.delete_record()
        elif 'submit' in request.vars:
            success, week, day, start = convert_date_to_weekdaytime(form.vars.start)
            form.vars.academic_week = week
            form.vars.day_of_week = day
            form.vars.start_time = start
            event_record.update_record(**form.vars)
        elif 'duplicate' in request.vars:
            db.events.insert(**db.events._filter_fields(event_record))
        
        # The redirect passes back the date of the record to allow the calendar
        # to stay on the same calendar pane.
        redirect(URL(args=[module_id], vars={'last_date': event_record.start.date()}))
    
    return dict(module_data=_module_page_header(module_id, 'Events'),
                event_data=event_data, form=form)

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def module_view():
    """A controller to combine module and event email and then display, converting
    markdown to HTML along the way. Anyone can view.
    """
    
    module_id = request.args[0] 
    
    content= module_markdown(module_id)
    
    return dict(content = XML(markdown.markdown(content, extensions=['def_list'])), 
                module_data=_module_page_header(module_id, 'HTML'))

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def module_doc():
    """A controller to push out a document version of the view, using pandoc to convert 
    markdown into docx or latex. Anyone can access.
    """
    
    module_id = request.args[0] 
    output_format = request.args[1] 
    
    content= module_markdown(module_id, title=True)
    
    if output_format == 'docx':
        # With docx, the output _has_ to be written to a file, not just handled internally
        # as a string. Also, oddly, streaming the response from an open file directly 
        # reuslt in weird stalling behaviour and failure, but loading it to bytes and 
        # streaming the content of that works flawlessly. <shrug>
        
        filename = f'Module_{module_id}_{datetime.date.today()}.docx'
        url = os.path.join('static', 'module_docx',  filename)
        filepath = os.path.join(request.folder, url)
        template = os.path.join(request.folder, 'static', 'module_docx',  'docx_template.docx')
        pypandoc.convert_text(content, 'docx', format='md',  outputfile=filepath,
                              extra_args=[f"--reference-doc={template}"])
    
        with open(filepath, 'rb') as fin:
            data = io.BytesIO(fin.read())
        ctype = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    
    elif output_format == 'latex':
        # With latex, can just directly pass the string
        filename = f'Module_{module_id}_{datetime.date.today()}.tex'
        data = pypandoc.convert_text(content, 'latex', format='md')
        ctype = 'application/x-tex'
    
    disposition = f'attachment; filename={filename}'
    raise HTTP(200, data,
               **{'Content-Type': ctype,
               'Content-Disposition': disposition})


## COURSE INFORMATION CONTROLLERS
## - courses/ shows a list of courses
## - courses/id shows a course level summary and links to modules
## - doc downloads the existing information as docx or latex

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def courses():
    """A controller to view the modules in a course
    """
    
    # update the format of teaching_staff rows for either response
    db.teaching_staff._format = lambda row: f" {row.first_name} {row.last_name}" 
    
    # if the URL has no arguments, present a list of courses with URLs
    if len(request.args) == 0:
        
        courses = db(db.courses).select()
        courses = list(courses.render())
        
        table_rows = []
        for crs in courses:
            row = TR(TD(B(crs.fullname), DIV(_style='flex:1'), 
                        A('View', _href=URL('courses', args=crs.id)),
                     _style='display:flex'))
            table_rows.append(row)
            
            if crs.coconvenor is None:
                conv_string = crs.convenor
            else:
                conv_string = f"{crs.convenor} & {crs.coconvenor}"
            row = TR(TD(DIV(_style='flex:1'), conv_string,
                     _style='display:flex'))
            table_rows.append(row)
            
            table = TABLE(table_rows, _class='table-striped')
        
        return dict(mod_table=table, series_table=DIV(), header=H2('Course list'))
    
    # Otherwise get the requested course module list
    course_id = int(request.args[0])
    
    course = db(db.courses.id == course_id).select()
    course = list(course.render())[0]

    modules = db((db.modules.courses.contains(course_id)) &
                 (db.modules.is_series != True)
                 ).select(db.modules.id, 
                          db.modules.title,
                          db.modules.convenors,
                          db.modules.placeholder_week,
                          db.modules.placeholder_n_weeks,)
    
    series = db((db.modules.courses.contains(course_id)) &
                 (db.modules.is_series == True)
                 ).select(db.modules.id, 
                          db.modules.title,
                          db.modules.convenors,
                          db.modules.placeholder_week,
                          db.modules.placeholder_n_weeks,)
    
    header = CAT(H2(course.fullname),
                 P(B('Course Convenor: '), course.convenor),
                 P(B('Course Co-convenor: '), course.coconvenor),
                 P(B('Download full timetable with events: '), 
                     SPAN(A('Word', _href=URL('course_doc', args=[course.id, 'docx', 1])),
                          _style="margin:5px"),
                     XML('&or;'),
                     SPAN(A('LaTeX', _href=URL('course_doc', args=[course.id, 'latex', 1])),
                          _style="margin:5px"),
                     ),
                P(B('Download module level timetable: '), 
                    SPAN(A('Word', _href=URL('course_doc', args=[course.id, 'docx', 0])),
                         _style="margin:5px"),
                    XML('&or;'),
                    SPAN(A('LaTeX', _href=URL('course_doc', args=[course.id, 'latex', 0])),
                         _style="margin:5px"),
                    ))
    
    # Process any series modules
    if series:
        series = list(series.render())
        
        # Add dates and sort in time order
        _ = [update_module_record_with_dates(row) for row in series]
        series.sort(key=lambda row: row.start)
    
        # Process into a table - would be nice to still use fullCalendar but
        # it doesn't provide a week slot only day and time - could use the 
        # same hack as for the module grid, but that doesn't seem worth it
        
        series_table_rows = []
        for mod in series:
            row = TR(TD(B(mod.title), 
                        DIV(_style='flex:1'), 
                        A('View', _href=URL('module_view', args=mod.id)),
                        _style='display:flex'))
            series_table_rows.append(row)
            row = TR(TD(f'{mod.start} to {mod.end}',
                        DIV(_style='flex:1'), 
                        mod.convenors,
                        _style='display:flex'))
            series_table_rows.append(row)
            
        series_table = TABLE(series_table_rows, _class='table-striped', _style='width:100%')
        series_table = CAT(H3('Series modules'), series_table)
    else:
        series_table = DIV()
    
    # Now process standard modules
    # Convert ids to representation
    modules = list(modules.render())
    
    # Add dates and sort in time order
    _ = [update_module_record_with_dates(row) for row in modules]
    modules.sort(key=lambda row: row.start)
    
    # Process into a table - would be nice to still use fullCalendar but
    # it doesn't provide a week slot only day and time
    mod_table_rows = []
    for mod in modules:
        week = ((mod.start - FIRST_DAY).days // 7) + 1
        row = TR(TD(f'Week {week}', _style='width:20%'),
                 TD(B(mod.title), DIV(_style='flex:1'), 
                    A('View', _href=URL('module_view', args=mod.id)),
                 _style='display:flex'))
        mod_table_rows.append(row)
        row = TR(TD(),
                 TD(mod.start, DIV(_style='flex:1'), mod.convenors,
                 _style='display:flex'))
        mod_table_rows.append(row)
    
    mod_table = TABLE(mod_table_rows, _class='table-striped', _style='width:100%')
    mod_table = CAT(H3('Module list'), mod_table)
    
    return dict(header=header, 
                mod_table=mod_table,
                series_table=series_table)


@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def course_doc():
    """A controller to push out a DOCX version of a course, using pandoc to convert 
    markdown into the docx. Anyone can access.
    """
    
    course_id = int(request.args[0])
    output_format = request.args[1]
    show_events = bool(int(request.args[2]))
    
    content = ""
    
    # update the format of teaching_staff rows
    db.teaching_staff._format = lambda row: f" {row.first_name} {row.last_name}" 
    
    course = db.courses[course_id]
    
    ## STANDARD MODULE DATA AND SUMMARY TABLE
    modules = db(db.modules.courses.contains(course_id) &
                 (db.modules.is_series != True)
                 ).select(db.modules.id,
                          db.modules.title,
                          db.modules.convenors,
                          db.modules.placeholder_week,
                          db.modules.placeholder_n_weeks)
    
    # Convert ids to representation
    modules = list(modules.render())
    
    # Add dates and sort in time order
    _ = [update_module_record_with_dates(row) for row in modules]
    modules.sort(key=lambda row: row.start)
    
    # Process into a simple pipe table
    content += '## Weekly module summary\n\n\n'
    
    table = '\n\nWeek|Starting|Module|Convenors\n'
    table += '-----|-----|-----|-----\n'
    for mod in modules:
        week = ((mod.start - FIRST_DAY).days // 7) + 1
        table += f'{week}|{mod.start.strftime("%-d %b %Y")}|{mod.title}|{mod.convenors}\n'
    
    content += table
    
    ## SERIES MODULE DATA AND SUMMARY TABLE
    series = db(db.modules.courses.contains(course_id) &
                 (db.modules.is_series == True)
                 ).select(db.modules.id,
                          db.modules.title,
                          db.modules.convenors,
                          db.modules.placeholder_week,
                          db.modules.placeholder_n_weeks)
    
    if series:
        # Convert ids to representation
        series = list(series.render())
    
        # Add dates and sort in time order
        _ = [update_module_record_with_dates(row) for row in series]
        series.sort(key=lambda row: row.start)
    
        # Process into a simple pipe table
        content += '## Series module summary\n\n\n'
    
        table = '\n\nStart|End|Module|Convenors\n'
        table += '-----|-----|-----|-----\n'
        for mod in series:
            table += f'{mod.start.strftime("%-d %b %Y")}|{mod.end.strftime("%-d %b %Y")}|{mod.title}|{mod.convenors}\n'
    
        content += table
    
    ## STANDARD MODULE DETAILS
    content += '\n\n## Module details\n\n'
    
    for mod in modules:
        content += module_markdown(mod.id, title=True, show_events=show_events)
    
    if series:
        content += '\n\n## Series module details\n\n'
    
        for mod in series:
            content += module_markdown(mod.id, title=True, show_events=show_events)
    
    if output_format == 'docx':
        # With docx, the output _has_ to be written to a file, not just handled internally
        # as a string. Also, oddly, streaming the response from an open file directly 
        # reuslt in weird stalling behaviour and failure, but loading it to bytes and 
        # streaming the content of that works flawlessly. <shrug>
        
        filename = f'{course.abbrname}_{datetime.date.today()}.docx'
        url = os.path.join('static', 'module_docx',  filename)
        filepath = os.path.join(request.folder, url)
        template = os.path.join(request.folder, 'static', 'module_docx',  'docx_template.docx')
        pypandoc.convert_text(content, 'docx', format='markdown+pipe_tables', 
                              outputfile=filepath,
                              extra_args=[f"--reference-doc={template}"])
    
        with open(filepath, 'rb') as fin:
            data = io.BytesIO(fin.read())
        ctype = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    
    elif output_format == 'latex':
        # With latex, can just directly pass the string
        filename = f'{course.abbrname}_{datetime.date.today()}.tex'
        data = pypandoc.convert_text(content, 'latex', format='markdown+pipe_tables')
        ctype = 'application/x-tex'
        
    disposition = f'attachment; filename={filename}'
    raise HTTP(200, data,
               **{'Content-Type': ctype,
               'Content-Disposition': disposition})

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def integrated_timetable():
    """Creates a selectable list of courses to use in all_modules_doc"""

    db.courses.id.readable = False

    form =   SQLFORM.grid(
                db.courses,
                fields=[db.courses.fullname, db.courses.id],
                selectable= [
                    ("Download DOCX",
                    lambda ids : redirect(
                        URL('timetabler', 'all_modules_doc', 
                            args=['_'.join([str(i) for i in ids]), 'docx', '1']))
                    ),
                    ("Download Latex",
                    lambda ids : redirect(
                        URL('timetabler', 'all_modules_doc', 
                            args=['_'.join([str(i) for i in ids]), 'latex', '1']))
                    ),
                    ("Download Markdown",
                    lambda ids : redirect(
                        URL('timetabler', 'all_modules_doc', 
                            args=['_'.join([str(i) for i in ids]), 'md', '1']))
                    )
                ],
                details=False,
                create=False,
                deletable=False,
                editable=False,
                csv=False,
                maxtextlength=100
    )

    return dict(form = form)


@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def all_modules_doc():
    """A controller to push out a DOCX version of all modules in the timetabler, across
    a set of courses, using pandoc to convert markdown into the docx. Anyone can access.
    """

    # Something subsitutes commas for _ in parsed request.args
    course_ids = [int(v) for v in request.args[0].split('_')]
    output_format = request.args[1]
    show_events = bool(int(request.args[2]))

    content = "## Course listing\n\n"

    # update the format of teaching_staff rows
    db.teaching_staff._format = lambda row: f" {row.first_name} {row.last_name}" 

    # Create a course table
    courses_data = db(
        (db.courses.id.belongs(course_ids)) &
        (db.courses.convenor == db.teaching_staff.id)
        ).select(orderby=db.courses.fullname)

    table = '\nCourse|Convenor|Module link\n'
    table += '-----|-----|-----|-----\n'

    for crs in courses_data.render():
        table += (f"{crs.courses.fullname}|"
                  f"{crs.teaching_staff.first_name} {crs.teaching_staff.last_name}|"
                  f"[Modules](#{IS_SLUG()(crs.courses.fullname)[0]})\n")

    content += table

    # Create course specific module summaries
    content += "\n\n## Module overviews by course\n\n"

    for crs in courses_data:

        content += f"### {crs.courses.fullname}\n\n"

        ## STANDARD MODULE DATA AND SUMMARY TABLE
        modules = db(db.modules.courses.contains(crs.courses.id) &
                    (db.modules.is_series != True)
                    ).select(db.modules.id,
                            db.modules.title,
                            db.modules.convenors,
                            db.modules.placeholder_week,
                            db.modules.placeholder_n_weeks)
        
        # Convert ids to representation
        modules = list(modules.render())
        
        # Add dates and sort in time order
        _ = [update_module_record_with_dates(row) for row in modules]
        modules.sort(key=lambda row: row.start)
        
        # Process into a simple pipe table
        content += f'#### Standard module summary\n\n'
        
        table = 'Week|Starting|Module|Convenors\n'
        table += '-----|-----|-----|-----\n'
        for mod in modules:
            week = ((mod.start - FIRST_DAY).days // 7) + 1
            slug = IS_SLUG()(mod.title)[0]
            table += f'{week}|{mod.start.strftime("%-d %b %Y")}|[{mod.title}](#{slug})|{mod.convenors}\n'
        
        content += table
        
        ## SERIES MODULE DATA AND SUMMARY TABLE
        series = db(db.modules.courses.contains(crs.courses.id) &
                    (db.modules.is_series == True)
                    ).select(db.modules.id,
                            db.modules.title,
                            db.modules.convenors,
                            db.modules.placeholder_week,
                            db.modules.placeholder_n_weeks)
        
        if series:
            # Convert ids to representation
            series = list(series.render())
        
            # Add dates and sort in time order
            _ = [update_module_record_with_dates(row) for row in series]
            series.sort(key=lambda row: row.start)
        
            # Process into a simple pipe table
            content += f'\n\n#### Series module summary\n\n'
        
            table = 'Start|End|Module|Convenors\n'
            table += '-----|-----|-----|-----\n'
            for mod in series:
                slug = IS_SLUG()(mod.title)[0]
                table += f'{mod.start.strftime("%-d %b %Y")}|{mod.end.strftime("%-d %b %Y")}|[{mod.title}](#{slug})|{mod.convenors}\n'

            content += table
    
    ## STANDARD MODULE DETAILS ACROSS COURSES
    content += '\n\n## Module listing\n\n'
    
    # The table definitions are unfortunate here - should have used normalised course to
    # module table. There are two lists - the list[int] field from db.modules and a list
    # of courses via request.args[0]. We want to known if _any_ values from the first
    # list occur in the second. 
    #
    # * .contains(val) finds if a single course is in the list[int] field
    # * .belongs(list) does the opposite, single value in list[int] occurs in args.
    #
    # So - post filter using sets.

    all_modules = db(db.modules
                ).select(db.modules.id,
                        db.modules.title,
                        db.modules.convenors,
                        db.modules.placeholder_week,
                        db.modules.placeholder_n_weeks,
                        db.modules.is_series,
                        db.modules.courses,
                        orderby=db.modules.placeholder_week)

    
    modules  = [m for m in all_modules 
                if set(m.courses).intersection(course_ids) and not m.is_series]

    series = [m for m in all_modules 
                if set(m.courses).intersection(course_ids) and m.is_series] 


    for mod in modules:
        content += module_markdown(mod.id, title=True, show_events=show_events)

    if series:
        content += '\n\n## Series module details\n\n'
    
        for mod in series:
            content += module_markdown(mod.id, title=True, show_events=show_events)
    
    if output_format == 'docx':
        # With docx, the output _has_ to be written to a file, not just handled internally
        # as a string. Also, oddly, streaming the response from an open file directly 
        # reuslt in weird stalling behaviour and failure, but loading it to bytes and 
        # streaming the content of that works flawlessly. <shrug>
        
        filename = f'Module_download_{datetime.date.today()}.docx'
        url = os.path.join('static', 'module_docx',  filename)
        filepath = os.path.join(request.folder, url)
        template = os.path.join(request.folder, 'static', 'module_docx',  'docx_template.docx')
        pypandoc.convert_text(content, 'docx', format='markdown+pipe_tables', 
                              outputfile=filepath,
                              extra_args=[f"--reference-doc={template}"])
    
        with open(filepath, 'rb') as fin:
            data = io.BytesIO(fin.read())
        ctype = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    
    elif output_format == 'latex':
        # With latex, can just directly pass the string
        filename = f'Module_download_{datetime.date.today()}.tex'
        data = pypandoc.convert_text(content, 'latex', format='markdown+pipe_tables')
        ctype = 'application/x-tex'
    elif output_format == 'md':
        # With markdown, return content without modification
        filename = f'Module_download_{datetime.date.today()}.md'
        data = content
        ctype = 'text/markdown'
    else:
        raise HTTP(406, "Unknown output format")
        
    disposition = f'attachment; filename={filename}'
    raise HTTP(200, data,
               **{'Content-Type': ctype,
               'Content-Disposition': disposition})

## Other views

@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def module_grid():
    """A controller to expose the sequence of modules by course. Anyone can access.
    All of the actual work is done using clientside JS and JSON feeds.
    
    TODO - add admin level edit to move modules on calendar.
    """
    
    # Use a div to pass the year academic start data to the client for use by fullcalendar
    year_data = DIV(_id='year_data', _day_one=FIRST_DAY)
    
    return dict(year_data=year_data)



@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def room_grid():
    """A controller to view the weekly use of rooms by module. Anyone can access.
    """
    # Use a div to pass the year academic start data to the client for use by fullcalendar
    year_data = DIV(_id='year_data', _day_one=FIRST_DAY)
    
    return dict(year_data=year_data)


@auth.requires(
    auth.has_membership(auth.id_group("admin"))
    or auth.has_membership(auth.id_group("timetabler"))
)
def download_module_grid():
    
    """
    Converts the module grid into an Excel spreadsheet. This is not a great
    format for overlapping calendar events, but seems to be a pragmatic choice
    for a grid for playing with the layout. A calendar application would be 
    better but this is quicker to implement.
    """
    
    # Extract the data for modules
    modules = db((db.modules.is_series == False)
                 ).select(db.modules.id,  
                          db.modules.title, 
                          db.modules.courses, 
                          db.modules.is_series, 
                          db.modules.placeholder_week, 
                          db.modules.placeholder_n_weeks,
                          orderby=db.modules.placeholder_week)
    
    # Map course ids to names (might be a way to do this in the query)
    courses = db(db.courses).select(db.courses.id, db.courses.abbrname)
    course_map = {rw.id: rw.abbrname for rw in courses} 
    
    # Pack the modules into courses
    packing = {rw.id: [] for rw in courses}
    
    for ky in packing:
        packing[ky] = {wk: [] for wk in range(1,53)}
    
    # Share colours across modules
    colours = itertools.cycle(['8dd3c7','ffffb3','bebada','fb8072','80b1d3',
                               'fdb462','b3de69','fccde5','d9d9d9'])
    
    for mod in modules:
        
        wk_range = range(mod.placeholder_week, 
                         mod.placeholder_week + mod.placeholder_n_weeks)
        
        # Pack modules into weeks. There can be multiple modules per week,
        # so a list is used to hold the modules. Modules can also span weeks 
        # and these columns will be merged (within courses). To do this,
        # the first week will contain a tuple (name, n_weeks, col) and then a 
        # placeholder of None is used to 'hold' that column open in the list
        
        # Also need to ensure that the module is always in the same index - 
        # you can't just append as modules can run past the end of a blocking
        # module in an earlier index and then drop back to a lower index.
        
        col = next(colours)
        
        for crs in mod.courses:
            for week in wk_range:
                
                if week == mod.placeholder_week:
                    mod_idx = len(packing[crs][week])
                    packing[crs][week].append((mod.title, mod.placeholder_n_weeks, col))
                else:
                    if not packing[crs][week]:
                        packing[crs][week] = [None] * (mod_idx + 1)
                    else:
                        packing[crs][week].append(None)
    
    # Find the most modules in one week
    max_mod = max((len(packing[c][w]) for c in packing for w in range(1,53)))
    
    # Now create the workbook instance to be populated
    wb = openpyxl.Workbook()
    hdrfont = openpyxl.styles.Font(bold=True)
    mod_align = openpyxl.styles.Alignment(horizontal='general',
                                          vertical='top',
                                          text_rotation=0,
                                          wrap_text=True,
                                          shrink_to_fit=False,
                                          indent=0)
    
    # name the worksheet and add a heading
    ws = wb.active
    ws.title = 'Module Grid'
    ws.cell(row=1, column=1, value= f'Module grid downloaded {datetime.date.today().isoformat()}')
    ws.cell(row=1, column=1).font = hdrfont
    
    # Header row
    hdr_row = 3
    
    # Add week numbers
    cell = ws.cell(row=hdr_row, column=1)
    cell.value = 'Week'
    cell.font = hdrfont
    
    for week in range(1,52):
        cell = ws.cell(row=hdr_row + week, column=1)
        cell.value = week
        cell.font = hdrfont
    
    # Add course data
    current_left_col = 2
    
    for crs_id, crs_weeks in packing.items():
        
        cell = ws.cell(row=hdr_row, column=current_left_col)
        cell.value = course_map[crs_id]
        cell.font = hdrfont
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        ws.merge_cells(start_row=hdr_row, start_column=current_left_col, 
                       end_row=hdr_row, end_column=current_left_col + max_mod - 1)
        
        for this_week, these_modules in crs_weeks.items():
            
            for mod_idx, mod_data in enumerate(these_modules):
                
                # Check the this module -  None placeholder or (name, n_weeks, col) for merged cells
                if mod_data is None:
                    pass
                else:
                    # unpack the components
                    mod_name, n_weeks, mod_col = mod_data
                    
                    cell = ws.cell(row=hdr_row + this_week, column=current_left_col + mod_idx)
                    cell.value = mod_name
                    cell.alignment = mod_align
                    cell.fill = openpyxl.styles.fills.PatternFill("solid", fgColor=mod_col)
                    
                    if n_weeks > 1:
                        
                        ws.merge_cells(start_row=hdr_row + this_week, 
                                       start_column=current_left_col + mod_idx, 
                                       end_row=hdr_row + this_week + n_weeks - 1, 
                                       end_column=current_left_col + mod_idx)
        
        current_left_col += max_mod
    
    # Sizing of cells
    for col in range(2, len(course_map) * max_mod + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Prep the payload to be returned
    attachment = 'attachment;filename=Module_Grid_{}.xlsx'.format(datetime.date.today().isoformat())
    
    with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            
            raise HTTP(200, tmp.read(),
                       **{'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                          'Content-Disposition':attachment + ';'})


## Admin


@auth.requires_membership('admin')
def freezer():
    
    # By default, the registration key is not visible or editable
    rec = db(db.freezer).select().first()
    form = SQLFORM(db.freezer,
                   record=rec.id)
    
    if form.process().accepted:
        redirect(URL('index'))
    
    return dict(form = form)


##
## Data services
##
## fullcalendar automatically sends start and end to do lazy loading of events, 
## but that is currently not implemented in any of these services. Many of
## these services include an unused timeZone argument. It is passed in by fullcalendar
## but we aren't using it.


def call():
    session.forget()
    return service()

@service.json
def get_college_dates(start=None, end=None, timeZone=None):
    """Service to provide college dates as events, merging college dates and 
    recurring events into a single feed
    """
    
    college_dates = db(db.college_dates).select()
    
    json = []
    
    for record in college_dates:
        if record.event_startdate is not None:
            
            entry = dict(id=record.id,
                         title=record.name)
            if record.event_enddate is not None:
                entry['start'] = record.event_startdate
                entry['end'] = record.event_enddate
                entry['allDay'] = True
            elif record.event_duration is not None:
                start_time = datetime.time(9,0) if record.event_starttime is None else record.event_starttime
                entry['start'] = datetime.datetime.combine(record.event_startdate, start_time)
                entry['end'] = entry['start'] + datetime.timedelta(hours=record.event_duration)
                entry['allDay'] = False
            
            json.append(entry)
    
    recurring = db(db.recurring_events).select()
    
    for record in recurring:
        entry = dict({'daysOfWeek':[record.day_of_week],
                      'title':record.title,
                      'id':record.id, 
                      'startRecur': record.recur_startdate,
                      'endRecur': record.recur_enddate})
        if record.all_day:
            entry['allDay'] = True
        else:
            entry['startTime'] = record.start_time
            entry['endTime'] = record.end_time
        
        json.append(entry)
    
    return json


@service.json
def get_events(module_id, start=None, end=None, event_id=None, timeZone=None):
    """Service to provide the events within a module with locations as resources"""
    
    # Use the events set from the module to get the module details
    # and the events. Logged in users get to see concealed events
    if auth.has_membership('timetabler'):
        db.events._common_filter = None
    
    module = db.modules[module_id]
    events = module.events.select()
    
    events_json = []
    
    for ev in events:
        update_event_record_with_dates(ev)
        ev_dict = dict(id=ev.id,
                       title=ev.title,
                       start=ev.start,
                       end=ev.end,
                       color='cornflowerblue',
                       editable=False,
                       resourceIds=ev.location_id,
                       extendedProps=dict(description=ev.description,
                                          teacher_id=ev.teacher_id),
                       url=URL('module_events',args=[module_id, ev.id]))
        
        if event_id != 'null':
            event_id = int(event_id)
            
            if ev.conceal == True:
                ev_dict['color'] = 'grey'
            
            if ev.id == event_id:
                ev_dict['color'] = 'salmon'
            
            if ev.id == event_id and auth.has_membership('timetabler'):
                ev_dict['editable'] = True
        
        events_json.append(ev_dict)
        
    return events_json


@auth.requires_membership('timetabler')
@service.json
def post_new_event(module_id, datetime, duration):
    """Used to create a new event when fullcalendar fires eventReceive"""
    
    success, *data = convert_date_to_weekdaytime(datetime)
    
    if not success:
        return None
    
    recid = db.events.insert(title='New event',
                             academic_week=data[0],
                             day_of_week=data[1],
                             start_time=data[2],
                             duration=duration,
                             module_id=module_id)
    
    record = db.events[recid]

    return record.id


@service.json
def get_events_by_week(start=None, end=None, timeZone=None):
    """Service to provide a week of events with locations as resources"""
    
    # Get modules that intersect the start date
    start = parser.isoparse(start).date()
    modules = db(db.modules).select()
    [update_module_record_with_dates(mod) for mod in modules]
    
    these_modules = [mod for mod in modules if (mod.start <= start) and (mod.end > start)]
    
    if not these_modules:
        return []
    
    module_start_date = these_modules[0].start
    events = db(db.events.module_id.belongs([mod.id for mod in modules])).select()
    
    events_json = []
    
    for ev in events:
        update_event_record_with_dates(ev, module_start_date)
        
        ev = dict(id=ev.id,
                  title=ev.title,
                  start=ev.start,
                  end=ev.end,
                  resourceIds=ev.location_id,
                  extendedProps=dict(description=ev.description,
                                     teacher_id=ev.teacher_id),
                  url=URL('module_events',args=[ev.module_id, ev.id]))
        
        ev['color'] = 'grey'
        
        events_json.append(ev)
        
    return events_json


@service.json
def get_locations():
    """Service to provide locations as fullcalendar resource data"""
    
    locs = db(db.locations).select(db.locations.id, db.locations.title, db.locations.is_external)
    
    return locs


@service.json
def get_courses():
    """Service to provide courses as fullcalendar resource data"""
    
    courses = db(db.courses).select(db.courses.id, db.courses.abbrname)
    
    for crs in courses:
        crs.title = crs.abbrname
        crs.pop('abbrname')
    
    return courses


@service.json
def get_modules_old(start=None, end=None, course_id=None):
    """Service to provides module level events for the module grid. 
    
    This mechanism provides a sideways scrolling grid - it uses the resourceTimelineMonth
    view out of the box as intended, but having to click through months and splitting
    across month boundaries is clumsy and unpleasant to use. The new mechanism hacks a
    dayGrid to present a vertical scrolling view.
    """
    
    if course_id is None:
        query = (db.modules.is_series == False)
    else:
        query = ((db.modules.courses.contains(course_id)) & (db.modules.is_series == 0))
    
    modules = db(query).select(db.modules.id, 
                               db.modules.title,
                               db.modules.courses,
                               db.modules.is_series,
                               db.modules.placeholder_week,
                               db.modules.placeholder_n_weeks)
    
    # This is a bit clumsy - need to add start, end and url
    # and convert courses entry to resourceIDs
    _ = [update_module_record_with_dates(m, for_fullcalendar=True) for m in modules]
    for mod in modules:
        mod.url = URL('module_view', args=mod.id)
        mod.resourceIds = mod.courses
        mod.pop('courses')
    
    return modules

@service.json
def get_modules(start=None, end=None, course_id=None, timeZone=None):
    """Service to provides module level events for the module grid. 
    
    Fullcalendar does a great job of laying out events and providing structures
    but has some odd limitations in the default views. This mechanism is a horrible
    hack to get a vertical resource view with weekly slots. The modules are converted
    into 20 minute slots from 01:00 to 17:00 (48 slots) and passed out to a dayGrid,
    which then uses JS to relabel the hour slots as academic calendar weeks"""
    
    if course_id is None:
        query = (db.modules.is_series == False)
    else:
        query = ((db.modules.courses.contains(course_id)) & (db.modules.is_series == 0))
    
    modules = db(query).select(db.modules.id, 
                               db.modules.title,
                               db.modules.courses,
                               db.modules.is_series,
                               db.modules.placeholder_week,
                               db.modules.placeholder_n_weeks)
    
    # A repeating set containing a color brewer palette for giving each event different colours
    colours = itertools.cycle(['#8dd3c7','#ffffb3','#bebada','#fb8072','#80b1d3',
                               '#fdb462','#b3de69','#fccde5','#d9d9d9'])
    
    # This is a bit clumsy - need to add start, end and url
    # and convert courses entry to resourceIDs
    _ = [update_module_record_with_dates(m, for_fullcalendar=True) for m in modules]
    modules.sort(lambda x: x.start)
    
    for mod in modules:
        mod.url = URL('module_view', args=mod.id)
        mod.resourceIds = mod.courses
        midnight = datetime.datetime.combine(datetime.date.today(), datetime.time.min)
        mod.start = midnight + datetime.timedelta(minutes=(mod.placeholder_week) * 20 + 40)
        mod.end = mod.start + datetime.timedelta(minutes= 20) * mod.placeholder_n_weeks
        mod['backgroundColor'] = next(colours)
        mod.pop('courses')
    
    return modules
