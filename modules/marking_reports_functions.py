import datetime
import os
import sys
import csv
import itertools
import zipfile
import io
import re
import openpyxl
import simplejson as json
from collections import Counter
import fpdf
import importlib
from tempfile import NamedTemporaryFile
from html.parser import HTMLParser 
from mailer import Mail

from gluon import (current, SQLFORM, DIV, LABEL, CAT, B, P, A,
                   URL, HTTP, BR, TABLE, H2, H4, XML, Field, IS_NULL_OR, IS_IN_SET)

"""
This module contains key functions for processing marking reports. They have been 
separated out into a module to reduce the amount of code in the controller, which
should speed up response times.
"""


## --------------------------------------------------------------------------------
## GLOBAL FUNCTIONS
## --------------------------------------------------------------------------------

def div_radio_widget(field, value, **attributes):
    # provides a horizontal radio button rubric
    table=SQLFORM.widgets.radio.widget(field, value, **attributes)
    
    return DIV(*[DIV(DIV(LABEL(td.components[1], _style='display:block;'),
                         td.element('input'), ),
                         _class='col-sm-2', _style='display:inline-block;text-align:center;')
                 for td in table.elements('td') 
                     if '_disabled' not in list(td.element('input').attributes.keys())],
               _class='row')

def div_checkbox_widget(field, value, **attributes):
    # provides a horizontal checkbox rubric for three fields
    # - needs to catch hidden field at end
    table = SQLFORM.widgets.checkboxes.widget(field, value, **attributes)

    return DIV(*[DIV(DIV(td.element('input'), '  ',
                         td.element('label')),
                         _class='col-sm-3', _style='display:inline-block;')
                 for td in table.elements('td') 
                     if '_disabled' not in list(td.element('input').attributes.keys())],
               _class='row')

def div_checkbox_widget_wide(field, value, **attributes):
    # provides a horizontal checkbox rubric for three fields
    # - needs to catch hidden field at end
    table = SQLFORM.widgets.checkboxes.widget(field, value, **attributes)

    return DIV(*[DIV(DIV(td.element('input'), '  ',
                         td.element('label')),
                         _class='col-sm-6', _style='display:inline-block;')
                 for td in table.elements('td') 
                     if '_disabled' not in list(td.element('input').attributes.keys())],
               _class='row')

## --------------------------------------------------------------------------------
## LOCAL FUNCTIONS USED BY THE ASSIGNMENT PAGE TO TAKE ACTIONS ON SETS OF RECORD IDS
## --------------------------------------------------------------------------------

def release(ids):
    
    """
    Local function to email reports to students. It will only email released reports.
    """
    
    db = current.db
    
    # Get only submitted
    n=0
    email = []
    for id in ids:
        record = db.assignments(id)
        if record.status in ['Submitted', 'Released']:
            record.update_record(status='Released')
            email.append(record)
            n += 1
    
    # now group by student
    email.sort(key=lambda rec: rec['student_email'])
    student_blocks = {k:list(recs) 
                      for k, recs 
                      in itertools.groupby(email, lambda x: x['student_email'])}

    # Create a mailer instance to send multiple emails using the same connection
    fails = 0
    mailer = Mail()
    mailer.login()
    
    for student, recs in student_blocks.items():
        
        # Create a set of links to the reports
        links = [CAT(P(B('Marking Role: '), 
                       r.marker_role,
                       B('; Marker: '), 
                       A(r.marker.first_name, ' ', r.marker.last_name, 
                         _href=URL('download_pdf', scheme=True, host=True,
                                   vars={'record':r.id, 
                                         'public_access_token': r.public_access_token}))))
                for r in recs]
        
        success = mailer.sendmail(subject='Your Project Marking Reports', 
                                  to=student, email_template='student_release.html',
                                  email_template_dict={'name':recs[0].student_first_name + ' ' + recs[0].student_last_name,
                                                 'links':CAT(links).xml()})
        
        if not success:
            fails += 1
    
    mailer.logout()
    del mailer
    
    # give some feedback
    msg = 'Emailed {} released records from {} selected rows to {} students'.format(
            n, len(ids), len(student_blocks)-fails)
    if fails > 0:
        msg = msg + "Warning: FAILED to send {} emails. Review email log.".format(fails)
    
    current.session.flash = msg


def distribute(ids):
    
    """
    This local function emails a set of records out to the relevant markers
    Note that this can be repeated to send reminders but won't email submitted
    or released reports, to avoid hassling markers!
    """
    
    db = current.db
    
    # Update the database to show distribution: Created -> Not started
    qry_by_select_ids = db.assignments.id.belongs(ids)
    db(qry_by_select_ids & (db.assignments.status == 'Created')).update(status='Not started')
    
    # Find the uncompleted records and group by marker and role
    email = db(qry_by_select_ids & 
               (db.assignments.status.belongs(['Not started', 'Started'])) &
               (db.assignments.marker == db.markers.id)
               ).select(db.assignments.marker,
                        db.markers.id,
                        db.markers.email,
                        db.markers.first_name,
                        db.markers.marker_access_token,
                        db.assignments.marker_role_id,
                        db.assignments.marker_role_id.count().with_alias('n'),
                        groupby=(db.assignments.marker,
                                 db.markers.id,
                                 db.markers.email,
                                 db.markers.first_name,
                                 db.markers.marker_access_token,
                                 db.assignments.marker_role_id))
    
    # now render to text values and group by marker
    email = [e for e in email.render()]
    email.sort(key= lambda rec: rec['markers.email'])
    marker_blocks = {k:list(recs) 
                     for k, recs 
                     in itertools.groupby(email, lambda x: x['markers.email'])}
    
    # Create a mailer instance to send multiple emails using the same connection
    rec_fails = 0
    mail_fails = 0
    
    mailer = Mail()
    mailer.login()
    
    for marker, recs in marker_blocks.items():
        
        # now summarize the number of each type of report
        reports_to_submit = XML(TABLE([(r.assignments.marker_role_id, r.n) for r in recs]))
        
        # Create a link to the marker my assignments page.
        my_assignments_url = URL('my_assignments', scheme=True, host=True,
                                  vars={'marker': recs[0].markers.id, 
                                        'marker_access_token': recs[0].markers.marker_access_token})
        
        success = mailer.sendmail(subject='Silwood Park Masters Project Marking', 
                                  to=marker, email_template='marker_distribute.html',
                                  email_template_dict={'name':recs[0].markers.first_name,
                                                       'reports_to_submit': reports_to_submit,
                                                       'my_assignments_url': my_assignments_url})
        
        if not success:
            rec_fails += sum([r.n for r in recs])
            mail_fails += 1
    
    mailer.logout()
    del mailer
    
    # give some feedback
    n_row = len(ids)
    msg = f"Emailed {n_row - rec_fails} records from {n_row} selected rows to {len(marker_blocks)} markers."
    if mail_fails > 0:
        msg = msg + f" Warning: FAILED to send {mail_fails} emails. Review email log."
    
    current.session.flash = msg


def zip_pdfs(ids, confidential=False):
    
    """
    Local function to create a zipfile of pdfs for selected records and
    return it to the user
    """
    
    db = current.db
    
    # create a zipfile of the selected pdfs and return it
    contents = io.BytesIO()
    zf = zipfile.ZipFile(contents, mode='w', compression=zipfile.ZIP_DEFLATED)
    
    # loop over the ids, creating the correct PDF and adding it.
    for id in ids:
        
        record = db.assignments(id)
        
        # can't create a PDF of incomplete assignments
        if record.status in ['Submitted','Released']:
            
            # create the form
            pdf, filename = create_pdf(record, confidential=confidential)
            
            zf.writestr(filename, pdf)
    
    # finalise and return the zip file
    zf.close()
    
    current.response.headers['Content-Type'] = 'application/pdf'
    attachment = 'attachment;filename=MarkingRecords.zip'
    current.response.headers['Content-Disposition'] = attachment
    
    raise HTTP(200, contents.getvalue(),
               **{'Content-Type':'application/pdf',
                  'Content-Disposition':attachment + ';'})


def download_grades(ids):
    
    """
    This local function takes a set of rows from the assignments grid
    and returns a collated excel file, one row per student, with the 
    names and grades in separate columns. 
    """
    db = current.db
    wb = openpyxl.Workbook()
    
    # spreadsheet styles
    left = openpyxl.styles.Alignment(horizontal='left')
    center = openpyxl.styles.Alignment(horizontal='center')
    weekend = openpyxl.styles.PatternFill(fill_type='solid', start_color='CCCCCC')
    head = openpyxl.styles.Font(size=14, bold=True)
    subhead = openpyxl.styles.Font(bold=True)
    warn = openpyxl.styles.Font(bold=True, color='FF0000')
    cell_shade = {'Approved': openpyxl.styles.PatternFill(fill_type='solid', start_color='8CFF88'),
                  'Submitted': openpyxl.styles.PatternFill(fill_type='solid', start_color='FFCB8B'),
                  'Draft': openpyxl.styles.PatternFill(fill_type='solid', start_color='DDDDDD'),
                  'Rejected': openpyxl.styles.PatternFill(fill_type='solid', start_color='FF96C3')}

    # name the worksheet and add a heading
    ws = wb.active
    ws.title = 'Grades'
    ws['A1'] = 'Masters grades: downloaded {}'.format(datetime.datetime.today().isoformat())
    ws['A1'].font = head
    
    # load reports and then group reports by student and role
    records = db(db.assignments.id.belongs(ids)).select()
    records = list(records.render()) # to render marker id to name
    records.sort(key = lambda rw: (rw['student_cid'], rw['marker_role']))
    records = {key: list(group) for key, group 
                in itertools.groupby(records, lambda rw: (rw['student_cid'], rw['marker_role']))}
    
    # Find the unique student details, ordered by presentation and surname
    students = db(db.assignments.id.belongs(ids)).select(
                    db.assignments.student_cid,
                    db.assignments.student_last_name,
                    db.assignments.student_first_name,
                    db.assignments.course_presentation,
                    db.assignments.academic_year,
                    orderby=(db.assignments.course_presentation, 
                             db.assignments.student_last_name),
                    distinct=True)
    
    # Layout the headers and data above and below a fixed row
    data_row = 6
    hdr_row = data_row - 1
    
    ws[f'A{hdr_row}'] = 'CID'
    ws[f'B{hdr_row}'] = 'Last Name'
    ws[f'C{hdr_row}'] = 'First Name'
    ws[f'D{hdr_row}'] = 'Course Presentation'
    ws[f'E{hdr_row}'] = 'Year'
    
    # Convert students to a dictionary keyed by CID, assigning a data row per student
    student_dict = {}
    for row, this_student in enumerate(students):
        this_student['row'] = row + data_row
        student_dict[this_student.student_cid] = this_student
    
    # Fill in student details
    for cid, details in student_dict.items():
        
        ws.cell(row = details.row, column=1, value=details.student_cid)
        ws.cell(row = details.row, column=2, value=details.student_last_name)
        ws.cell(row = details.row, column=3, value=details.student_first_name)
        ws.cell(row = details.row, column=4, value=details.course_presentation)
        ws.cell(row = details.row, column=5, value=details.academic_year)
    
    # Now fill in report details for each kind of report in turn
    role_list = db(db.marking_roles).select()
    column_pointer = 6
    # A regex to convert '65% (B)' grades to numeric
    grade_regex = re.compile('[0-9]+(?=%)')
    
    for this_role in role_list:
        
        # Get the JSON data on what variables to export
        form_json = this_role.json
        grade_export = form_json['grade_export']
        cols_per_report = len(grade_export) + 1
        
        # Extract the columns
        filtered_to_role = {key[0]: val for key, val in records.items() if key[1] == this_role}
        
        # Loop over students and reports
        for this_student, these_reports in filtered_to_role.items():
            
            student_details = student_dict[this_student]
            row = student_details.row
            
            for report_num, this_report in enumerate(these_reports):
                
                report_data = this_report.get('assignment_data')
                
                for grade_num, this_grade in enumerate(grade_export):
                    
                    # Do we have any data for this grade - might be nothing at assignment level
                    # or might not be completed in the assignment data
                    val = None if report_data is None else report_data.get(this_grade)
                    val = 'NA' if val is None else val
                    
                    # look for a percentage grade
                    perc_grade = grade_regex.match(val)
                    val = val if perc_grade is None else int(perc_grade.group(0))
                    
                    ws.cell(row = row, 
                            column = column_pointer + cols_per_report * report_num + grade_num,
                            value = val)
                
                # Add marker
                ws.cell(row = row, 
                        column = column_pointer + cols_per_report * report_num + len(grade_export),
                        value = this_report.marker)
        
        # Add role headers
        max_n_reports = max([len(rep) for rep in filtered_to_role.values()])
        for report_number in range(max_n_reports):
            ws.cell(row = data_row - 2, 
                    column= column_pointer + report_number * cols_per_report,
                    value= f"{this_role} {report_number + 1}")
            
            for grade_num, this_grade in enumerate(grade_export):
                ws.cell(row = data_row - 1, 
                        column = column_pointer + cols_per_report * report_number + grade_num,
                        value = this_grade)
            
            ws.cell(row = data_row - 1, 
                    column= column_pointer + report_number * cols_per_report + cols_per_report - 1,
                    value='Marker')
        
        column_pointer += max_n_reports * cols_per_report
    
    current.response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    attachment = 'attachment;filename=Marking_Grades_{}.xlsx'.format(datetime.date.today().isoformat())
    current.response.headers['Content-Disposition'] = attachment
    
    with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            
            raise HTTP(200, tmp.read(),
                       **{'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                          'Content-Disposition':attachment + ';'})

## --------------------------------------------------------------------------------
## HTML Report writing functions shared by show_form and write_report
## --------------------------------------------------------------------------------

def get_form_header(form_json, record, readonly, security=None):
    """Takes a marking form definition and a marking assignment record
    and returns a standard header block for a form.
    """
    
    # Define the header block
    header_rows =   [('Student',  '{student_first_name} {student_last_name}'),
                     ('CID', '{student_cid}'),
                     ('Course Presentation', '{course_presentation_id}'),
                     ('Academic Year', '{academic_year}'),
                     ('Marker', '{marker}'),
                     ('Marker Role', '{marker_role_id}'),
                     ('Status', '{status}')]
     
    header = [DIV(LABEL(l, _class='col-sm-3'),
                  DIV(v.format(**record), _class='col-sm-9'),
                  _class='row')
              for l, v in header_rows]
    
    header.insert(0, H2(form_json['title']))
    header.append(DIV(LABEL('Marking critera', _class='col-sm-3'),
                      DIV(A(form_json['marking_criteria'],
                            _href=URL('static','marking_criteria/' + form_json['marking_criteria']),
                            _target='blank'),
                          _class='col-sm-9'),
                      _class='row'))
    
    # If readonly, provide a link to the PDF download link, otherwise, insert any instructions
    if readonly:
        header.append(H4('Completed report'))
        header.append(P('Click ', A('here', _href=URL('download_pdf', vars=security)),
                        ' to download a confidential pdf of the full report.'))
    else:
        header.append(H4('Instructions'))
        header.append(XML(form_json['instructions']))
    
    return header


def assignment_to_sqlform(record, readonly, buttons=[]):
    
    """Takes a marking assignment and generates the appropriate marking form,
    converting the fields defined in the form JSON description into a SQLFORM
    object. This form then can be processed for user input by the controller
    but is restyled for display using style_sqlform to modify the default
    SQLFORM representation.
    
    Returns the SQLFORM object and the form JSON description, which is used
    for validation.
    """
    
    # - Extract the set of fields from the components section of the questions array
    #   This allows a question to have multiple bits (e.g. rubric + optional comments)
    form_json = record.marker_role_id.form_json
    
    fields=[]
    for q in form_json['questions']:
        for c in q['components']:
            if c['type'] == 'rubric':
                fields.append(Field(c['variable'], 
                              type='string', 
                              requires=IS_NULL_OR(IS_IN_SET(c['options'])),
                              widget=div_radio_widget))
            elif c['type'] == 'comment':
                # protect carriage returns in the display of the text but stop anybody using 
                # any other tags
                fields.append(Field(c['variable'], 
                              type='text', 
                              represent=lambda text: "" if text is None else XML(text.replace('\n', '<br />'),
                              sanitize=True, 
                              permitted_tags=['br/'])))
            elif c['type'] == 'select':
                fields.append(Field(c['variable'], 
                              type='string', 
                              requires=IS_NULL_OR(IS_IN_SET(c['options']))))
            else:
                # Other types that don't insert a form control- currently only 'query'
                pass
    
    # - define the form for IO using the fields object,
    #   preloading any existing data
    if record.assignment_data in [None, '']:
        data_json=None
    else:
        data_json=record.assignment_data
    
    # - define a SQLFORM object the fields object for processing data
    form = SQLFORM.factory(*fields,
                           readonly=readonly,
                           buttons=buttons,
                           record=data_json,
                           showid=False)

    return form


def style_sqlform(record, form, readonly=False):
    
    form_json = record.marker_role_id.form_json
    
    # modify any widget settings for active forms
    if not readonly:
        for q in form_json['questions']:
            for c in q['components']:
                if 'nrow' in list(c.keys()):
                    form.custom.widget[c['variable']]['_rows'] = c['nrow']
                if 'placeholder' in list(c.keys()):
                    form.custom.widget[c['variable']].update(_placeholder=c['placeholder'])
                if 'value' in list(c.keys()):
                    form.custom.widget[c['variable']].components = [c['value']]
    
    # compile the laid out form in a list
    html = [form.custom.begin]
    
    # add the questions in the order they appear in the JSON array
    for q in form_json['questions']:
        html.append(DIV(H4(q['title']), _style='background-color:lightgrey;padding:1px'))
        
        if q.get('info') is not None:
            html.append(DIV(XML(q.get('info'))))
        else:
            html.append(DIV(_style='min-height:10px'))
        
        for c in q['components']:
            
            # Insert either the form variable and stored data or the output of a 
            # query component.
            if c['type'] == 'query':
                # This is a tricky line. globals() contains globals functions, so needs
                # the appropriate query function to be imported. All such functions
                # get the assignment record as an input. You could use eval() here but
                # that has security risks - hard to see them applying here, but hey.
                query_data = globals()[c['query']](record)
                html.append(DIV(DIV(B(c['label']), _class='col-sm-2'),
                                DIV(query_data, _class='col-sm-10'),
                                _class='row'))
            else:
                html.append(DIV(DIV(B(c['label']), _class='col-sm-2'),
                                DIV(form.custom.widget[c['variable']], _class='col-sm-10'),
                                _class='row'))
            
            if c.get('info') is not None:
                html.append(DIV(XML(c.get('info'))))
            else:
                html.append(DIV(_style='min-height:10px'))
            
        html.append(DIV(_style='min-height:10px'))
    
        
    # finalise the form and send the whole thing back
    html.append(BR())
    html.append(form.custom.submit)
    html.append(form.custom.end)
    
    return html


## --------------------------------------------------------------------------------
## PDF generation
## --------------------------------------------------------------------------------

# Subclass of FPDF with a built in confidential page header
class ConfidentialPDF(fpdf.FPDF):
    
    def header(self):
        self.set_y(7)
        self.set_font("Helvetica", size=12)
        self.set_text_color(255, 40, 40)
        self.cell(0, 8, txt = "This is a confidential document and "
                  "must not be circulated to students.", align='C')
        self.set_xy(10,20)


def create_pdf(record, confidential):
    
    """
    This code writes a simple PDF file of the marking report
    using pyfpdf: very simple, can't do styling or templating easily 
    but no external dependencies and reasonably fast
    """
    
    if confidential:
        pdf = ConfidentialPDF(format='A4')
    else:
        pdf = fpdf.FPDF(format='A4')
    
    
    form_json = record.marker_role_id.form_json
    
    # set up the fonts
    fpdf.set_global('SYSTEM_TTFONTS', current.configuration.get('fpdf.font_dir'))
    pdf.add_font('DejaVu', '', 'DejaVuSansCondensed.ttf', uni=True)
    pdf.add_font('DejaVu', 'B', 'DejaVuSansCondensed-Bold.ttf', uni=True)
    
    #  add first page and insert the logo
    pdf.set_top_margin(20)
    pdf.add_page()
    pdf.image(os.path.join(current.request.folder, 'static','images/imperial_logo_mono.png'), w=60)
    logo_bottom = pdf.get_y()

    # Title block
    pdf.set_xy(80, 20)
    pdf.set_font("DejaVu", style='B', size=14)
    pdf.cell(0, 10, txt="Silwood Park Masters Courses", align="R",  ln=1)
    pdf.cell(0, 10, txt=form_json['pdftitle'], align="R",  ln=1)
    pdf.set_xy(10, logo_bottom + 10)
    pdf.set_font("DejaVu", size=14)
    pdf.set_text_color(0, 0, 0)
    
    # ID table
    label =   ['Student', 'CID', 'Course', 'Year', 'Marker', 'Marker Role']
    content = ['{student_first_name} {student_last_name}',
               '{student_cid}',
               '{course_presentation}',
               '{academic_year}',
               '{marker.first_name} {marker.last_name}',
               '{marker_role}']
    
    for l, c in zip(label, content):
        pdf.set_font("DejaVu", size=12, style='B')
        pdf.cell(60, 8, txt=l, align="L")
        pdf.set_font("DejaVu", size=12)
        pdf.cell(0, 8, txt=c.format(**record), align="L", ln=1)
    
    
    # add the questions in the order they appear in the JSON array
    pdf.set_fill_color(200,200,200)
    
    for q in form_json['questions']:
        
        # skip confidential questions if the report is not confidential
        if not confidential and q['confidential']:
            continue
        
        # insert the question title in a grey bar
        pdf.rect(pdf.get_x(), pdf.get_y(), h=8, w=190, style='F')
        pdf.set_font("DejaVu", size=12, style='B')
        pdf.cell(0, 8, txt=q['title'], align="L", ln=1)
        
        # insert the components
        for c in q['components']:
            
            # skip confidential components if the report is not confidential
            if not confidential and q['confidential']:
                continue
            
            pdf.set_font("DejaVu", size=12, style='B')
            pdf.cell(60, 8, txt=c['label'], align="L")
            pdf.ln()
            pdf.set_font("DejaVu", size=12)
            
            if c['type'] == 'query':
                # This is a tricky line. globals() contains globals functions, so needs
                # the appropriate query function to be imported. All such functions
                # get the assignment record as an input. You could use eval() here but
                # that has security risks - hard to see them applying here, but hey.
                contents = globals()[c['query']](record, pdf=True)
            else:
                contents = record.assignment_data[c['variable']]
            
            if contents is None:
                contents = ""
            pdf.set_left_margin(70)
            pdf.write(6, txt=contents)
            pdf.set_left_margin(10)
            pdf.ln()
        
        # spacer
        pdf.ln()
    
    pdf.close()
    pdf = pdf.output(dest='S').encode('latin-1') # unicode string to bytes
    filename = '{} {} {} {} {} {}.pdf'.format(record.course_presentation,
                                            record.academic_year, 
                                            record.student_last_name,
                                            record.student_first_name,
                                            record.marker_role,
                                            record.id)
    
    return (pdf, filename)


## FORM QUERIES
## These function (currently only one) allow a form to include data by specifying a 'query'
## component in the form json. These should accept an assignments record as the first argument
## and then the option to provide either html (default, for web display) or a text repr
## for use in the PDF.

def query_report_marker_grades(record, pdf=False):
    """Creates a table of existing individual report marker grades
    """
    
    db = current.db
    
    # Bizarre issue - this is called with pdf=True from a page that has
    # altered the represent of status and that causes it to be masked from
    # the table, so reset that here to ensure that status can be read
    db.assignments.status.represent = None
    
    report_grades = db((db.assignments.student_cid == record.student_cid) &
                       (db.assignments.course_presentation == record.course_presentation) &
                       (db.assignments.academic_year == record.academic_year) &
                       (db.assignments.marker_role == 'Marker')
                       ).select(
                           db.assignments.marker,
                           db.assignments.assignment_data,
                           db.assignments.status)
    
    report_grades = list(report_grades.render())
    
    report_grades = [(rw.marker, rw.assignment_data['grade']) 
                        if rw.status in ['Submitted', 'Released']
                        else (rw.marker, 'Not submitted')
                        for rw in report_grades]
    
    if pdf:
        report_grades = [f'{rw[0]} ({rw[1]})' for rw in report_grades]
        return '\n'.join(report_grades)
    else:
        return TABLE(report_grades, _class='table')

## --------------------------------------------------------------------------------
## WIKI FUNCTIONS
## --------------------------------------------------------------------------------

class FoldingTOC(HTMLParser):
    """
    This parser scans html and inserts ids, classes and spans to turn
    <ul> or <ol> into a click to expand table of contents tree. The
    layout_wiki.html file contains JS and CSS to make it work.
    """
    
    def __init__(self):
        super().__init__()
        self.content = []
        self.depth = 0
        self.last_li_index = 0

    def handle_starttag(self, tag, attrs): 

        if tag in ['ul', 'ol']:
            if self.depth == 0:
                self.content.append(f'<{tag.upper()} id="root">')
            else:
                self.content.append(f'<{tag.upper()} class="nested">')
                self.content[self.last_li_index] = '<LI><SPAN class="caret"></SPAN>'
            self.depth += 1
        
        elif tag == 'li':
            self.content.append('<LI><SPAN class="end"></SPAN>')
            self.last_li_index = len(self.content) - 1
        
        else:
            self.content.append(self.get_starttag_text())
        
    def handle_data(self, data): 
        self.content.append(data)
    
    def handle_endtag(self, tag):
        if tag in ['ul', 'ol']:
            self.depth -= 1
        
        self.content.append(f'</{tag.upper()}>')
    
    def get_toc(self):
        
        return ''.join(self.content)