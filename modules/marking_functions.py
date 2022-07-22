import datetime
import os
import sys
import csv
import itertools
import zipfile
import io
import re
import openpyxl
import simplejson as json
from collections import Counter
from itertools import groupby
import fpdf
import importlib
from tempfile import NamedTemporaryFile
from mailer import Mail

from gluon import (current, SQLFORM, DIV, LABEL, CAT, B, P, A, SPAN, INPUT,
                   URL, HTTP, BR, TABLE, H2, H4, XML, Field, IS_NULL_OR, IS_IN_SET)

from gluon.sqlhtml import OptionsWidget

"""
This module contains key functions for processing marking reports. They have been 
separated out into a module to reduce the amount of code in the controller, which
should speed up response times.
"""


## --------------------------------------------------------------------------------
## GLOBAL FUNCTIONS
## --------------------------------------------------------------------------------



def get_project_rollover_date():
    """Get the project rollover date, defaulting to October 1st next year"""    
    db = current.db
    project_rollover = db(
        db.college_dates.name == 'Project Rollover Date'
        ).select().first()
    
    if project_rollover is not None:
        return project_rollover.event_startdate
    else:
        today = datetime.date.today()
        if today.month > 9:
            return datetime.date(today.year + 1, 10, 1)
        else:
            return datetime.date(today.year, 10, 1)


def div_radio_widget(field, value, **attributes):
    # provides a horizontal radio button rubric
    table=SQLFORM.widgets.radio.widget(field, value, **attributes)
    
    return DIV(*[DIV(DIV(LABEL(td.components[1], _style='display:block;'),
                         td.element('input'), ),
                         _class='col-sm-2', _style='display:inline-block;text-align:center;')
                 for td in table.elements('td') 
                     if '_disabled' not in list(td.element('input').attributes.keys())],
               _class='row')

def div_checkbox_widget(field, value, **attributes):
    # provides a horizontal checkbox rubric for three fields
    # - needs to catch hidden field at end
    table = SQLFORM.widgets.checkboxes.widget(field, value, **attributes)

    return DIV(*[DIV(DIV(td.element('input'), '  ',
                         td.element('label')),
                         _class='col-sm-3', _style='display:inline-block;')
                 for td in table.elements('td') 
                     if '_disabled' not in list(td.element('input').attributes.keys())],
               _class='row')

def div_checkbox_widget_wide(field, value, **attributes):
    # provides a horizontal checkbox rubric for three fields
    # - needs to catch hidden field at end
    table = SQLFORM.widgets.checkboxes.widget(field, value, **attributes)

    return DIV(*[DIV(DIV(td.element('input'), '  ',
                         td.element('label')),
                         _class='col-sm-6', _style='display:inline-block;')
                 for td in table.elements('td') 
                     if '_disabled' not in list(td.element('input').attributes.keys())],
               _class='row')

def div_checkbox_widget_list_group(field, value, **attributes):
    
    # This is just to repackage the checkbox widgets and allow some styling
    # (which should be done with CSS) - also adds a disabled input with the field
    # name, which is used for display of form.errors
    
    table = SQLFORM.widgets.checkboxes.widget(field, value, **attributes)
    
    return CAT(DIV(*[DIV(td.element('input'), SPAN(_style='padding:0px 5px;'),
                        LABEL(td.element('label').components[0],
                              _for=td.element('label').attributes['_for'],
                              _style='margin:0px;'),
                            _class='list-group-item',
                            _style='padding:0px 20px; background:lightgrey')
                    for td in table.elements('td')
                        if '_disabled' not in list(td.element('input').attributes.keys())],
                   _class='list-group', _name=field.name),
                INPUT(_style="display:none;",
                      _disabled="disabled",
                      _name=field.name,
                      hideerror=False))

## --------------------------------------------------------------------------------
## LOCAL FUNCTIONS USED BY THE ASSIGNMENT PAGE TO TAKE ACTIONS ON SETS OF RECORD IDS
## --------------------------------------------------------------------------------

def release(ids):
    
    """
    Local function to email reports to students. It will only email released reports.
    """
    
    db = current.db    
    
    qry_by_select_ids = ((db.assignments.id.belongs(ids)) &
                         (db.assignments.status.belongs(['Submitted', 'Released'])))
    
    db(qry_by_select_ids).update(status='Released')
    
    # Find students covered by these records
    email = db(qry_by_select_ids & 
               (db.assignments.student_presentation == db.student_presentations.id) &
               (db.student_presentations.student == db.students.id)
               ).select(db.students.student_email,
                        db.students.student_first_name,
                        db.students.student_last_name,
                        db.students.student_access_token,
                        db.assignments.marker,
                        db.assignments.marker_role_id,
                        db.assignments.id)
    
    # now group by student email (can't compare rows, so use unique email)
    # and then create groups of records using students detail tuples as keys
    email = [rec for rec in email.render()]
    email.sort(key = lambda rec: rec['students']['student_email'])
    student_blocks = {tuple(k.values()): list(recs) 
                      for k, recs 
                      in itertools.groupby(email, lambda x: (x['students']))}

    # Create a mailer instance to send multiple emails using the same connection
    fails = 0
    mailer = Mail()
    mailer.login()
    
    for (st_email, st_first, st_last, token), recs in student_blocks.items():
        
        # Create a set of links to the reports
        links = [CAT(P(B('Marking Role: '), 
                       r.assignments.marker_role_id,
                       B('; Marker: '), 
                       A(r.assignments.marker, 
                         _href=URL('download_pdf', scheme=True, host=True,
                                   vars={'record':r.assignments.id, 
                                         'student_access_token': token}))))
                for r in recs]
        
        success = mailer.sendmail(subject='Your Project Marking Reports', 
                                  to=st_email, 
                                  email_template='student_release.html',
                                  email_template_dict={'name':f"{st_first} {st_last}",
                                                       'links':CAT(links).xml()})
        
        if not success:
            fails += 1
    
    mailer.logout()
    del mailer
    
    # give some feedback
    msg = 'Emailed {} released records from {} selected rows to {} students'.format(
            len(email), len(ids), len(student_blocks) - fails)
    if fails > 0:
        msg = msg + "Warning: FAILED to send {} emails. Review email log.".format(fails)
    
    current.session.flash = msg


def distribute(ids):
    
    """
    This local function emails a set of records out to the relevant markers
    Note that this can be repeated to send reminders but won't email submitted
    or released reports, to avoid hassling markers!
    """
    
    db = current.db
    
    # Update the database to show distribution: Created -> Not started
    qry_by_select_ids = db.assignments.id.belongs(ids)
    db(qry_by_select_ids & (db.assignments.status == 'Created')).update(status='Not started')
    
    # Find the uncompleted records and group by marker and role
    email = db(qry_by_select_ids & 
               (db.assignments.status.belongs(['Not started', 'Started'])) &
               (db.assignments.marker == db.teaching_staff.id)
               ).select(db.assignments.marker,
                        db.teaching_staff.id,
                        db.teaching_staff.email,
                        db.teaching_staff.first_name,
                        db.assignments.marker_role_id,
                        db.assignments.marker_role_id.count().with_alias('n'),
                        groupby=(db.assignments.marker,
                                 db.teaching_staff.id,
                                 db.teaching_staff.email,
                                 db.teaching_staff.first_name,
                                 db.assignments.marker_role_id))
    
    # now render to text values and group by marker
    email = [e for e in email.render()]
    email.sort(key= lambda rec: rec['teaching_staff.email'])
    marker_blocks = {k:list(recs) 
                     for k, recs 
                     in itertools.groupby(email, lambda x: x['teaching_staff.email'])}
    
    # Create a mailer instance to send multiple emails using the same connection
    rec_fails = 0
    mail_fails = 0
    
    mailer = Mail()
    mailer.login()
    
    for marker, recs in marker_blocks.items():
        
        # now summarize the number of each type of report
        reports_to_submit = [(r.assignments.marker_role_id, r.n) for r in recs]
        
        # Create a link to the login URL
        login_url = URL('staff', 'staff_login', scheme=True, host=True,
                                  vars={'email': marker})
        
        success = mailer.sendmail(subject='Silwood Park Masters Project Marking', 
                                  to=marker, email_template='marker_distribute.html',
                                  email_template_dict={'name':recs[0].teaching_staff.first_name,
                                                       'reports_to_submit': reports_to_submit,
                                                       'login_url': login_url})
        
        if not success:
            rec_fails += sum([r.n for r in recs])
            mail_fails += 1
    
    mailer.logout()
    del mailer
    
    # give some feedback
    n_row = len(ids)
    msg = f"Emailed {n_row - rec_fails} records from {n_row} selected rows to {len(marker_blocks)} markers."
    if mail_fails > 0:
        msg = msg + f" Warning: FAILED to send {mail_fails} emails. Review email log."
    
    current.session.flash = msg


def zip_pdfs(ids, confidential=False):
    
    """
    Local function to create a zipfile of pdfs for selected records and
    return it to the user
    """
    
    db = current.db
    
    # create a zipfile of the selected pdfs and return it
    contents = io.BytesIO()
    zf = zipfile.ZipFile(contents, mode='w', compression=zipfile.ZIP_DEFLATED)
    
    # loop over the ids, creating the correct PDF and adding it.
    for id in ids:
        
        record = db.assignments(id)
        
        # can't create a PDF of incomplete assignments
        if record.status in ['Submitted','Released']:
            
            # create the form
            pdf, filename = create_pdf(record, confidential=confidential)
            
            zf.writestr(filename, pdf)
    
    # finalise and return the zip file
    zf.close()
    
    current.response.headers['Content-Type'] = 'application/pdf'
    attachment = 'attachment;filename=MarkingRecords.zip'
    current.response.headers['Content-Disposition'] = attachment
    
    raise HTTP(200, contents.getvalue(),
               **{'Content-Type':'application/pdf',
                  'Content-Disposition':attachment + ';'})


def download_grades(ids):
    
    """
    This local function takes a set of rows from the assignments grid
    and returns a collated excel file, one row per student, with the 
    names and grades in separate columns. 
    """
    db = current.db
    
    # Extract the data for the assignments to be collated
    records = db((db.assignments.id.belongs(ids)) &
                 (db.assignments.student_presentation == db.student_presentations.id) &
                 (db.student_presentations.student == db.students.id)
                 ).select(
                    db.students.student_cid,
                    db.students.student_last_name,
                    db.students.student_first_name,
                    db.student_presentations.course_presentation,
                    db.student_presentations.academic_year,
                    db.assignments.marker_role_id,
                    db.assignments.marker,
                    db.assignments.assignment_data,
                    orderby=(db.student_presentations.course_presentation,
                             db.students.student_cid))
    
    # Find the set of marking roles in the selected data and then 
    # get the set of exported details for each role
    roles = {r.assignments.marker_role_id for r in records}
    role_export = {}
    
    for this_role in roles:
        role_details = db.marking_roles[this_role]
        role_export[role_details.name] = role_details.form_json['grade_export']
    
    # Can now render the records to substitute marker names etc for id values
    records = list(records.render())
    
    # The next section figures out what columns are being written - the exported
    # values for possibly multiple instances of different roles.
    
    # 1) Get the count of each role by students and hence the maximum number of
    #    each role across the set of records - might be faster ways of doing this
    #    using sort and groupby, but this is shorter!
    student_role_count = Counter([(r.students.student_cid, r.assignments.marker_role_id) for r in records])
    
    role_max = {ky:0 for ky in role_export}
    for (_, rid), val in student_role_count.items():
        role_max[rid] = max(role_max[rid], val)
    
    # 2) Now map export details to columns by marker role. This map is structured
    # as a list of tuples with the following structure:
    #    [(role_name, role_instance_name, marker_column, 
    #      ((export field, column number), ...), 
    #      ...]
    # When the assignments for a given student are processed, a copy of the map is taken
    # and then the role ids for the student can be used to pop off a map entry with
    # matching role id for each assignment in turn
    
    data_col = 6
    field_map = []
    
    for this_role, this_max in role_max.items():
        # get the role export details
        these_fields = role_export[this_role]
        n_fields = len(these_fields)
        
        for this_copy in range(this_max):
            # get the tuple contents and increment the columns
            this_name = this_role + '_' + str(this_copy + 1)
            marker_col = data_col
            these_columns = list(range(data_col + 1, data_col + 1 + n_fields))
            data_col += n_fields + 1
            field_map.append((this_role, this_name, marker_col, list(zip(these_fields, these_columns))))
    
    # Now create the workbook instance to be populated
    wb = openpyxl.Workbook()
    hdrfont = openpyxl.styles.Font(bold=True)

    # name the worksheet and add a heading
    ws = wb.active
    ws.title = 'Grades'
    ws.cell(row=1, column=1, value= f'Grades downloaded {datetime.datetime.today().isoformat()}')
    ws.cell(row=1, column=1).font = hdrfont
    
    # Layout the standard headers
    data_row = 6
    hdr_row = data_row - 1  # Contains the field name
    role_row = data_row - 2  # Contains role header to discriminate fields
    
    std_headers = [(1, 'CID'),
                   (2, 'Last Name'),
                   (3, 'First Name'),
                   (4, 'Course Presentation'),
                   (5, 'Year')]
    
    for col, hdr in std_headers:
        cell = ws.cell(row=hdr_row, column=col)
        cell.value = hdr
        cell.font = hdrfont
    
    # Write the field map structure to the header rows
    for _, role, marker, fields in field_map:
        # insert the role header
        cell = ws.cell(row = role_row, column = marker)
        cell.value = role
        cell.font = hdrfont
        
        # Marker column
        cell = ws.cell(row = hdr_row, column = marker)
        cell.value = 'marker'
        cell.font = hdrfont
        
        # insert the field headers
        for hdr, col in fields:
            cell = ws.cell(row = hdr_row, column = col)
            cell.value = hdr
            cell.font = hdrfont
    
    # A regex to convert '65% (B)' grades to numeric
    grade_regex = re.compile('[0-9]+(?=%)')
    
    # Now group records by student and export to rows
    grouped_records = groupby(records, key=lambda row: row.students.student_cid)
    
    for rw_idx, (this_student, student_records) in enumerate(grouped_records):
        
        # Populate a list of the records
        student_records = list(student_records)
        
        this_row = data_row + rw_idx
        
        # write out the student details from the first entry
        first = student_records[0]
        ws.cell(row=this_row, column=1, value=first.students.student_cid)
        ws.cell(row=this_row, column=2, value=first.students.student_last_name)
        ws.cell(row=this_row, column=3, value=first.students.student_first_name)
        ws.cell(row=this_row, column=4, value=first.student_presentations.course_presentation)
        ws.cell(row=this_row, column=5, value=first.student_presentations.academic_year)
        
        # Copy the field map for this student
        this_map = field_map.copy()
        
        # Loop over the student records
        for rec in student_records:
            
            # Pop a role slot from the map for the record
            role_slot_index = [mp[0] for mp in this_map].index(rec.assignments.marker_role_id)
            role_slot = this_map.pop(role_slot_index)
            
            # put in the marker
            ws.cell(row = this_row, column = role_slot[2], value = rec.assignments.marker)
            
            # get the form data
            report_data = rec.get('assignments.assignment_data')
            
            # Fill in the fields to be exported
            for fld, col in role_slot[3]:
                
                # Do we have any data for this grade - might be nothing at assignment level
                # or might not be completed in the assignment data
                val = None if report_data is None else report_data.get(fld)
                val = 'NA' if val is None else val
                
                # look for a percentage grade
                perc_grade = grade_regex.match(val)
                val = val if perc_grade is None else int(perc_grade.group(0))
                
                ws.cell(row = this_row, column = col, value = val)
    
    current.response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    attachment = 'attachment;filename=Marking_Grades_{}.xlsx'.format(datetime.date.today().isoformat())
    current.response.headers['Content-Disposition'] = attachment
    
    with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            
            raise HTTP(200, tmp.read(),
                       **{'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                          'Content-Disposition':attachment + ';'})

## --------------------------------------------------------------------------------
## HTML Report writing functions shared by show_form and write_report
## --------------------------------------------------------------------------------

def get_form_header(record, readonly, security=None, show_due_date=False):
    """Takes a marking form definition and a marking assignment record
    and returns a standard header block for a form.
    """
    
    # Define the header block
    header_rows =   [('Student',  '{student_presentation.student.student_first_name} {student_presentation.student.student_last_name}'),
                     ('CID', '{student_presentation.student.student_cid:09d}'),
                     ('Course Presentation', '{student_presentation.course_presentation.name}'),
                     ('Academic Year', '{student_presentation.academic_year}'),
                     ('Marker', '{marker.first_name} {marker.last_name}'),
                     ('Marker Role', '{marker_role_id.name}'),
                     ('Status', '{status}')]
    
    if show_due_date:
        header_rows.append(('Due Date', '{due_date}'))
        
    header = [DIV(LABEL(l, _class='col-sm-3'),
                  DIV(v.format(**record), _class='col-sm-9'),
                  _class='row')
              for l, v in header_rows]
    
    form_json = record.marker_role_id.form_json
    
    header.insert(0, H2(form_json['title']))
    header.append(DIV(LABEL('Marking critera', _class='col-sm-3'),
                      DIV(A(form_json['marking_criteria'],
                            _href=URL('static','marking_criteria/' + form_json['marking_criteria']),
                            _target='blank'),
                          _class='col-sm-9'),
                      _class='row'))
    
    # If readonly, provide a link to the PDF download link, otherwise, insert any instructions
    if readonly:
        header.append(H4('Completed report'))
        header.append(P('Click ', A('here', _href=URL('download_pdf', vars=security)),
                        ' to download a confidential pdf of the full report.'))
    else:
        header.append(H4('Instructions'))
        header.append(XML(form_json['instructions']))
    
    return header


def assignment_to_sqlform(record, readonly, buttons=[]):
    
    """Takes a marking assignment and generates the appropriate marking form,
    converting the fields defined in the form JSON description into a SQLFORM
    object. This form then can be processed for user input by the controller
    but is restyled for display using style_sqlform to modify the default
    SQLFORM representation.
    
    Returns the SQLFORM object and the form JSON description, which is used
    for validation.
    """
    
    # - Extract the set of fields from the components section of the questions array
    #   This allows a question to have multiple bits (e.g. rubric + optional comments)
    form_json = record.marker_role_id.form_json
    
    fields=[]
    for q in form_json['questions']:
        for c in q['components']:
            if c['type'] == 'rubric':
                fields.append(Field(c['variable'], 
                              type='string', 
                              requires=IS_NULL_OR(IS_IN_SET(c['options'])),
                              widget=div_radio_widget))
            elif c['type'] == 'comment':
                # protect carriage returns in the display of the text but stop anybody using 
                # any other tags
                fields.append(Field(c['variable'], 
                              type='text', 
                              represent=lambda text: "" if text is None else XML(text.replace('\n', '<br />'),
                              sanitize=True, 
                              permitted_tags=['br/'])))
            elif c['type'] == 'select':
                fields.append(Field(c['variable'], 
                              type='string', 
                              requires=IS_NULL_OR(IS_IN_SET(c['options']))))
            else:
                # Other types that don't insert a form control- currently only 'query'
                pass
    
    # - define the form for IO using the fields object,
    #   preloading any existing data
    if record.assignment_data in [None, '']:
        data_json=None
    else:
        data_json=record.assignment_data
    
    # - define a SQLFORM object the fields object for processing data
    form = SQLFORM.factory(*fields,
                           readonly=readonly,
                           buttons=buttons,
                           record=data_json,
                           showid=False)

    return form


def style_sqlform(record, form, readonly=False):
    
    form_json = record.marker_role_id.form_json
    
    # modify any widget settings for active forms
    if not readonly:
        for q in form_json['questions']:
            for c in q['components']:
                if 'nrow' in list(c.keys()):
                    form.custom.widget[c['variable']]['_rows'] = c['nrow']
                if 'placeholder' in list(c.keys()):
                    form.custom.widget[c['variable']].update(_placeholder=c['placeholder'])
                if 'value' in list(c.keys()):
                    form.custom.widget[c['variable']].components = [c['value']]
    
    # compile the laid out form in a list
    html = [form.custom.begin]
    
    # add the questions in the order they appear in the JSON array
    for q in form_json['questions']:
        html.append(DIV(H4(q['title']), _style='background-color:lightgrey;padding:1px'))
        
        if q.get('info') is not None:
            html.append(DIV(XML(q.get('info'))))
        else:
            html.append(DIV(_style='min-height:10px'))
        
        for c in q['components']:
            
            # Insert either the form variable and stored data or the output of a 
            # query component.
            if c['type'] == 'query':
                # This is a tricky line. globals() contains globals functions, so needs
                # the appropriate query function to be imported. All such functions
                # get the assignment record as an input. You could use eval() here but
                # that has security risks - hard to see them applying here, but hey.
                query_data = globals()[c['query']](record)
                html.append(DIV(DIV(B(c['label']), _class='col-sm-2'),
                                DIV(query_data, _class='col-sm-10'),
                                _class='row'))
            else:
                html.append(DIV(DIV(B(c['label']), _class='col-sm-2'),
                                DIV(form.custom.widget[c['variable']], _class='col-sm-10'),
                                _class='row'))
            
            if c.get('info') is not None:
                html.append(DIV(XML(c.get('info'))))
            else:
                html.append(DIV(_style='min-height:10px'))
            
        html.append(DIV(_style='min-height:10px'))
    
        
    # finalise the form and send the whole thing back
    html.append(BR())
    html.append(form.custom.submit)
    html.append(form.custom.end)
    
    return html


## --------------------------------------------------------------------------------
## PDF generation
## --------------------------------------------------------------------------------

# Subclass of FPDF with a built in confidential page header
class ConfidentialPDF(fpdf.FPDF):
    
    def header(self):
        self.set_y(7)
        self.set_font("Helvetica", size=12)
        self.set_text_color(255, 40, 40)
        self.cell(0, 8, txt = "This is a confidential document and "
                  "must not be circulated to students.", align='C')
        self.set_xy(10,20)


def create_pdf(record, confidential):
    
    """
    This code writes a simple PDF file of the marking report
    using pyfpdf: very simple, can't do styling or templating easily 
    but no external dependencies and reasonably fast
    """
    
    if confidential:
        pdf = ConfidentialPDF(format='A4')
    else:
        pdf = fpdf.FPDF(format='A4')
    
    
    form_json = record.marker_role_id.form_json
    
    # set up the fonts
    fpdf.set_global('SYSTEM_TTFONTS', current.configuration.get('fpdf.font_dir'))
    pdf.add_font('DejaVu', '', 'DejaVuSansCondensed.ttf', uni=True)
    pdf.add_font('DejaVu', 'B', 'DejaVuSansCondensed-Bold.ttf', uni=True)
    
    #  add first page and insert the logo
    pdf.set_top_margin(20)
    pdf.add_page()
    pdf.image(os.path.join(current.request.folder, 'static','images/imperial_logo_mono.png'), w=60)
    logo_bottom = pdf.get_y()

    # Title block
    pdf.set_xy(80, 20)
    pdf.set_font("DejaVu", style='B', size=14)
    pdf.cell(0, 10, txt="Silwood Park Masters Courses", align="R",  ln=1)
    pdf.cell(0, 10, txt=form_json['pdftitle'], align="R",  ln=1)
    pdf.set_xy(10, logo_bottom + 10)
    pdf.set_font("DejaVu", size=14)
    pdf.set_text_color(0, 0, 0)
    
    # ID table
    label =   ['Student', 'CID', 'Course Presentation', 'Year', 'Marker', 'Marker Role']
    content = ['{student_presentation.student.student_first_name} '  # Intentionally no comma
               '{student_presentation.student.student_last_name}',
               '{student_presentation.student.student_cid}',
               '{student_presentation.course_presentation.name}',
               '{student_presentation.academic_year}',
               '{marker.first_name} {marker.last_name}',
               '{marker_role_id.name}']
    
    for l, c in zip(label, content):
        pdf.set_font("DejaVu", size=12, style='B')
        pdf.cell(60, 8, txt=l, align="L")
        pdf.set_font("DejaVu", size=12)
        pdf.cell(0, 8, txt=c.format(**record), align="L", ln=1)
    
    
    # add the questions in the order they appear in the JSON array
    pdf.set_fill_color(200,200,200)
    
    for q in form_json['questions']:
        
        # skip confidential questions if the report is not confidential
        if not confidential and q['confidential']:
            continue
        
        # insert the question title in a grey bar
        pdf.rect(pdf.get_x(), pdf.get_y(), h=8, w=190, style='F')
        pdf.set_font("DejaVu", size=12, style='B')
        pdf.cell(0, 8, txt=q['title'], align="L", ln=1)
        
        # insert the components
        for c in q['components']:
            
            # skip confidential components if the report is not confidential
            if not confidential and q['confidential']:
                continue
            
            pdf.set_font("DejaVu", size=12, style='B')
            pdf.cell(60, 8, txt=c['label'], align="L")
            pdf.ln()
            pdf.set_font("DejaVu", size=12)
            
            if c['type'] == 'query':
                # This is a tricky line. globals() contains globals functions, so needs
                # the appropriate query function to be imported. All such functions
                # get the assignment record as an input. You could use eval() here but
                # that has security risks - hard to see them applying here, but hey.
                contents = globals()[c['query']](record, pdf=True)
            else:
                contents = record.assignment_data[c['variable']]
            
            if contents is None:
                contents = ""
            pdf.set_left_margin(70)
            pdf.write(6, txt=contents)
            pdf.set_left_margin(10)
            pdf.ln()
        
        # spacer
        pdf.ln()
    
    pdf.close()
    pdf = pdf.output(dest='S').encode('latin-1') # unicode string to bytes
    filename = '{} {} {} {} {} {}.pdf'.format(record.student_presentation.course_presentation.name,
                                            record.student_presentation.academic_year, 
                                            record.student_presentation.student.student_last_name,
                                            record.student_presentation.student.student_first_name,
                                            record.marker_role_id.name,
                                            record.id)
    
    return (pdf, filename)


## FORM QUERIES
## These function (currently only one) allow a form to include data by specifying a 'query'
## component in the form json. These should accept an assignments record as the first argument
## and then the option to provide either html (default, for web display) or a text repr
## for use in the PDF.

def query_report_marker_grades(record, pdf=False):
    """Creates a table of existing individual report marker grades
    """
    
    db = current.db
    
    # Bizarre issue - this is called with pdf=True from a page that has
    # altered the represent of status and that causes it to be masked from
    # the table, so reset that here to ensure that status can be read
    db.assignments.status.represent = None
    
    report_grades = db(#(db.student_presentations.id == record.student_presentation) &
                       (db.assignments.student_presentation == record.student_presentation) &
                       (db.assignments.marker_role_id == db.marking_roles.id) &
                       (db.marking_roles.name== 'Marker')
                       ).select(
                           db.assignments.marker,
                           db.assignments.assignment_data,
                           db.assignments.status)
    
    report_grades = list(report_grades.render())
    
    report_grades = [(rw.marker, rw.assignment_data['grade']) 
                        if rw.status in ['Submitted', 'Released']
                        else (rw.marker, 'Not submitted')
                        for rw in report_grades]
    
    if pdf:
        report_grades = [f'{rw[0]} ({rw[1]})' for rw in report_grades]
        return '\n'.join(report_grades)
    else:
        return TABLE(report_grades, _class='table')

